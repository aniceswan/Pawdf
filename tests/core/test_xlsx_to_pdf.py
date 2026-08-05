import datetime

import pytest

from pawdf.core._shared import ConversionError, page_count
from pawdf.core.xlsx_to_pdf import MAX_CELLS, xlsx_to_pdf


def _text_of(path, index=0):
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument(path)
    try:
        textpage = doc[index].get_textpage()
        try:
            return textpage.get_text_range()
        finally:
            textpage.close()
    finally:
        doc.close()


@pytest.fixture
def workbook(tmp_path):
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"
    sheet.append(["Region", "Q1", "Q2"])
    sheet.append(["North", 100, 150])
    sheet.append(["South", 90, 120])

    second = book.create_sheet("Notes")
    second.append(["Date", datetime.date(2026, 3, 1)])
    second.append(["Approved", True])
    second.append(["Ratio", 3.0])

    path = tmp_path / "book.xlsx"
    book.save(path)
    return path


def test_one_page_per_sheet(workbook, tmp_path):
    out = xlsx_to_pdf(workbook, tmp_path / "out.pdf")
    assert page_count(out) == 2


def test_cell_values_reach_the_page(workbook, tmp_path):
    out = xlsx_to_pdf(workbook, tmp_path / "out.pdf")
    text = _text_of(out, 0)

    assert "Sales" in text
    for value in ("Region", "North", "150"):
        assert value in text


@pytest.mark.parametrize(
    ("value", "expected"),
    [(datetime.date(2026, 3, 1), "2026-03-01"), (True, "TRUE"), (3.0, "3")],
)
def test_values_are_formatted_for_reading(tmp_path, value, expected):
    from openpyxl import Workbook

    book = Workbook()
    book.active.append(["label", value])
    path = tmp_path / "one.xlsx"
    book.save(path)

    out = xlsx_to_pdf(path, tmp_path / "out.pdf")
    assert expected in _text_of(out)


def test_whole_floats_lose_their_decimal(tmp_path):
    """A count column reading 3.0 instead of 3 looks like a bug to a reader."""
    from openpyxl import Workbook

    book = Workbook()
    book.active.append(["n", 3.0])
    path = tmp_path / "one.xlsx"
    book.save(path)

    assert "3.0" not in _text_of(xlsx_to_pdf(path, tmp_path / "out.pdf"))


def test_a_formatted_but_empty_sheet_does_not_explode(tmp_path):
    """openpyxl's max_row counts cells that were merely visited, so trusting
    it turns a two-row table into a hundred empty pages.
    """
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.append(["only", "row"])
    # Merely reaching for a cell creates it, which is what inflates max_row.
    sheet.cell(row=500, column=8)
    path = tmp_path / "sparse.xlsx"
    book.save(path)

    out = xlsx_to_pdf(path, tmp_path / "out.pdf")
    assert page_count(out) == 1


def test_an_oversized_sheet_is_truncated_and_says_so(tmp_path):
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    for row in range(MAX_CELLS + 200):
        sheet.append([f"r{row}"])
    path = tmp_path / "big.xlsx"
    book.save(path)

    out = xlsx_to_pdf(path, tmp_path / "out.pdf")
    combined = "".join(_text_of(out, i) for i in range(page_count(out)))
    assert "truncated" in combined.lower()


def test_an_empty_sheet_is_labelled(tmp_path):
    from openpyxl import Workbook

    book = Workbook()
    path = tmp_path / "empty.xlsx"
    book.save(path)

    out = xlsx_to_pdf(path, tmp_path / "out.pdf")
    assert "empty sheet" in _text_of(out)


def test_unreadable_input_raises(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not a workbook")

    with pytest.raises(ConversionError):
        xlsx_to_pdf(bad, tmp_path / "out.pdf")


def test_missing_input_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        xlsx_to_pdf(tmp_path / "nope.xlsx", tmp_path / "out.pdf")
