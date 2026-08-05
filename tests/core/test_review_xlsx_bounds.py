"""Regression tests for bounded Excel dimension discovery."""

from __future__ import annotations

from collections import Counter

import pytest

from pawdf.core._shared import ConversionError, page_count


def test_far_corner_formatting_does_not_scan_excel_maximum_grid():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    from pawdf.core.xlsx_to_pdf import _used_bounds

    book = Workbook()
    sheet = book.active
    sheet.append(["only", "row"])
    sheet.cell(row=1_048_576, column=16_384).fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )

    assert _used_bounds(sheet) == (1, 2)


def test_used_bounds_is_computed_once_per_visible_sheet(
    tmp_path,
    monkeypatch,
):
    from openpyxl import Workbook

    import pawdf.core.xlsx_to_pdf as converter

    book = Workbook()
    book.active.title = "One"
    book.active.append(["a", "b"])
    book.create_sheet("Two").append(["c", "d"])
    source = tmp_path / "cached-bounds.xlsx"
    book.save(source)

    calls: list[str] = []
    original = converter._used_bounds

    def counted(sheet):
        calls.append(sheet.title)
        return original(sheet)

    monkeypatch.setattr(converter, "_used_bounds", counted)
    output = converter.xlsx_to_pdf(source, tmp_path / "cached-bounds.pdf")

    assert page_count(output) == 2
    assert Counter(calls) == Counter({"One": 1, "Two": 1})


def test_excessive_materialized_cells_are_rejected_before_scanning():
    from pawdf.core.xlsx_to_pdf import (
        MAX_BOUND_SCAN_CELLS,
        _used_bounds,
    )

    class TooManyCells(dict):
        def __len__(self):
            return MAX_BOUND_SCAN_CELLS + 1

        def values(self):
            raise AssertionError("oversized cell map must not be scanned")

    class FakeSheet:
        title = "Oversized"
        _cells = TooManyCells()

    with pytest.raises(ConversionError, match="too many materialized cells"):
        _used_bounds(FakeSheet())
