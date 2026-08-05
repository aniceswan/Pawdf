"""Excel (.xlsx) -> PDF, without Excel or LibreOffice.

openpyxl reads the workbook, reportlab draws it. Best-effort by design, in the
same spirit as the Word converters: the point is that the app depends on
nothing but Python, not that it reproduces Excel's rendering.

What that costs is stated plainly in docs/conversion_fidelity.md. The short
version: values, not formulas' live behaviour; a grid, not Excel's layout
engine; no charts, no images, no conditional formatting.

Standalone feature package. See README.md in this directory.
"""

from __future__ import annotations

from pathlib import Path

from pawdf.core._shared import ConversionError, ensure_exists, ensure_parent_dir

__all__ = ["MAX_BOUND_SCAN_CELLS", "MAX_CELLS", "xlsx_to_pdf"]

#: A hard ceiling on cells rendered per sheet. A spreadsheet with a hundred
#: thousand rows is not a document, and trying to lay one out as a table means
#: reportlab building a flowable per cell until memory runs out. Truncating and
#: saying so is better than dying halfway through.
MAX_CELLS = 20_000
MAX_COLUMNS = 40
MAX_BOUND_SCAN_CELLS = MAX_CELLS * 10

#: Column width in Excel's units is roughly a character; this converts to points.
_WIDTH_UNIT_PT = 6.0
_MIN_COL_PT = 24.0
_MAX_COL_PT = 190.0


def _cell_text(value: object) -> str:
    """A cell as text, formatted the way a reader expects to see it."""
    import datetime

    if value is None:
        return ""
    if isinstance(value, bool):
        # Checked before int, because bool is an int in Python and "True"
        # reads better than "1" for a checkbox column.
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.datetime):
        return (
            value.strftime("%Y-%m-%d %H:%M")
            if any((value.hour, value.minute))
            else (value.strftime("%Y-%m-%d"))
        )
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _used_bounds(sheet) -> tuple[int, int]:  # noqa: ANN001 - openpyxl Worksheet
    """Return real value bounds without walking Excel's reported grid.

    Excel can report its full 1,048,576 by 16,384 grid after a user formats a
    distant cell or an entire sheet. ``iter_rows()`` trusts those dimensions,
    so using it here can perform billions of empty visits before Pawdf's
    rendering limit is applied. A normal openpyxl ``Worksheet`` already keeps
    only materialized cells in ``_cells``; inspecting that bounded mapping is
    exact for values and independent of inflated sheet dimensions.
    """

    cells = getattr(sheet, "_cells", None)
    if not isinstance(cells, dict):
        raise ConversionError(f"Worksheet '{sheet.title}' does not expose a safe cell map.")
    if len(cells) > MAX_BOUND_SCAN_CELLS:
        raise ConversionError(
            f"Worksheet '{sheet.title}' has too many materialized cells "
            f"to inspect safely ({len(cells):,}; limit "
            f"{MAX_BOUND_SCAN_CELLS:,})."
        )

    last_row = 0
    last_col = 0
    for cell in cells.values():
        if cell.value not in (None, ""):
            last_row = max(last_row, cell.row)
            last_col = max(last_col, cell.column)
    return last_row, last_col


def _column_widths(sheet, columns: int, available_pt: float) -> list[float]:  # noqa: ANN001
    """Column widths in points, honouring the sheet's own where it has them."""
    from openpyxl.utils import get_column_letter

    widths = []
    for index in range(1, columns + 1):
        dimension = sheet.column_dimensions.get(get_column_letter(index))
        raw = getattr(dimension, "width", None) if dimension else None
        points = (raw or 12) * _WIDTH_UNIT_PT
        widths.append(min(max(points, _MIN_COL_PT), _MAX_COL_PT))

    total = sum(widths)
    if total > available_pt:
        # Scale down to fit rather than letting reportlab clip the last
        # columns off the page silently.
        scale = available_pt / total
        widths = [w * scale for w in widths]
    return widths


def xlsx_to_pdf(
    input_path: str | Path,
    out_path: str | Path,
    *,
    landscape_wide_sheets: bool = True,
    show_gridlines: bool = True,
) -> Path:
    """Convert a workbook to PDF, one section per worksheet.

    `landscape_wide_sheets` turns a sheet that is wider than it is tall onto
    its side, which is what makes a broad table readable at all.
    """
    from openpyxl import load_workbook
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    p = ensure_exists(input_path)
    out_path = ensure_parent_dir(out_path)

    try:
        # data_only: read the last computed values, not the formula text.
        # Without it every calculated cell renders as "=SUM(A1:A9)".
        workbook = load_workbook(p, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001 - normalized for callers
        raise ConversionError(f"Could not read '{p}': {exc}") from exc

    title_style = ParagraphStyle(
        "SheetTitle", fontName="Helvetica-Bold", fontSize=14, leading=18, spaceAfter=10
    )
    cell_style = ParagraphStyle("Cell", fontName="Helvetica", fontSize=8, leading=10)
    head_style = ParagraphStyle("Head", fontName="Helvetica-Bold", fontSize=8, leading=10)
    note_style = ParagraphStyle(
        "Note", fontName="Helvetica-Oblique", fontSize=8, leading=11, textColor=colors.grey
    )

    try:
        sheets = [s for s in workbook.worksheets if s.sheet_state == "visible"]
        if not sheets:
            raise ConversionError("The workbook has no visible sheets.")

        # Compute each worksheet's true bounds once. The same cached result
        # chooses orientation and controls the rendered range.
        bounded_sheets = [(sheet, _used_bounds(sheet)) for sheet in sheets]

        # One page size for the document, chosen from the widest sheet: a
        # SimpleDocTemplate cannot change orientation partway through.
        widest = max(
            (bounds[1] for _sheet, bounds in bounded_sheets),
            default=0,
        )
        page_size = landscape(A4) if (landscape_wide_sheets and widest > 6) else A4
        doc = SimpleDocTemplate(
            str(out_path),
            pagesize=page_size,
            leftMargin=28,
            rightMargin=28,
            topMargin=28,
            bottomMargin=28,
        )
        available = page_size[0] - 56

        story: list = []
        for sheet_index, (sheet, bounds) in enumerate(bounded_sheets):
            if sheet_index > 0:
                story.append(PageBreak())
            story.append(Paragraph(sheet.title, title_style))

            rows, columns = bounds
            if not rows or not columns:
                story.append(Paragraph("(empty sheet)", note_style))
                continue

            truncated_columns = columns > MAX_COLUMNS
            columns = min(columns, MAX_COLUMNS)
            max_rows = max(1, MAX_CELLS // max(columns, 1))
            truncated_rows = rows > max_rows
            rows = min(rows, max_rows)

            data = []
            for row_index, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=rows, max_col=columns)
            ):
                style = head_style if row_index == 0 else cell_style
                data.append([Paragraph(_cell_text(c.value), style) for c in row])

            table = Table(data, colWidths=_column_widths(sheet, columns, available), repeatRows=1)
            grid_colour = colors.HexColor("#c8ccd4") if show_gridlines else colors.transparent
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.4, grid_colour),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef0f4")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                )
            )
            story.append(table)

            if truncated_rows or truncated_columns:
                story.append(Spacer(1, 8))
                story.append(
                    Paragraph(
                        "This sheet was larger than the export limit and has been "
                        f"truncated to {rows} rows and {columns} columns.",
                        note_style,
                    )
                )

        doc.build(story)
    except ConversionError:
        raise
    except Exception as exc:  # noqa: BLE001 - normalized for callers
        raise ConversionError(f"Could not convert the workbook to PDF: {exc}") from exc
    finally:
        workbook.close()

    return out_path
